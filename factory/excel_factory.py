from planilhas import ExcelRepository, AlterarExcelRepository
from usecase import UseCase, Funcionario
import openpyxl as op

class ExcelFactory:

    @staticmethod
    def create() -> UseCase:
        planilha = op.Workbook()
        pagina = planilha['Sheet']

        while True:
            informacoes = input('O que deseja adicionar?').split(';')
            pagina.append(informacoes)
            pergunta = input('Deseja continuar?').lower()

            if pergunta == 'n':
                break

        nome = input('Como vamos nomear sua planilha?')

        planilha.save(f'{nome}.xlsx')

        return UseCase(ExcelRepository.Create_sheet(planilha=planilha, informacoes=informacoes, nome=nome))
    
class AlterarFactory:

    @staticmethod
    def Alterar() -> Funcionario:
        try:    
            planilha = input('Qual planilha? ')
            planilha = op.load_workbook(f'{planilha}.xlsx')
            pagina = input('Qual pagina está o que deve ser alterado? ')
            pagina = planilha[f'{pagina}']
            alterar = input('O que devemos alterar nessa pagina? ')
            alteracao = input('Pelo que quer alterar? ')
            for rows in pagina.iter_rows(min_row=1, max_row=1000):
                for cell in rows:
                    if cell.value == alterar:
                        cell.value = alteracao
            save = input('Como deseja salvar? ')
            planilha.save(f'{save}.xlsx')
            return Funcionario(AlterarExcelRepository.Alter_sheet(planilha=planilha, pagina=pagina, substituicao=alterar, substituto=alteracao))
        except FileNotFoundError as error:
            print(f'{error}, Arquivo não encontrado')
        except KeyError:
            print('Por favor, verifique se os dados estão certos.')            